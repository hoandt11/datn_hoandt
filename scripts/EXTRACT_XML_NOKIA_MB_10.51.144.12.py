import xml.etree.ElementTree as ET
import csv
import configparser
from datetime import datetime
import os
import glob
import gzip
import shutil
from openpyxl import Workbook, load_workbook

def create_excel_log(log_file):
    """Tạo file Excel log nếu chưa tồn tại."""
    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Extraction Log"
        ws.append(["STT", "FOLDER_NAME", "FILE_NAME", "TIME_PROCESS", "FOLDER_NUMBER"])
        wb.save(log_file)

def append_to_excel_log(log_file, stt, folder_name, file_name, time_process, folder_number):
    """Thêm bản ghi vào file Excel log."""
    wb = load_workbook(log_file)
    ws = wb.active
    ws.append([stt, folder_name, file_name, time_process, folder_number])
    wb.save(log_file)

def get_next_folder_number(output_dir):
    """Tìm số thư mục lớn nhất trong output_dir và trả về số tiếp theo."""
    if not os.path.exists(output_dir):
        return 1
    folders = [f for f in os.listdir(output_dir) if os.path.isdir(os.path.join(output_dir, f)) and f.isdigit()]
    if not folders:
        return 1
    return max(int(f) for f in folders) + 1

def parse_conf_file(conf_file):
    """Đọc file .conf và trả về dictionary chứa measurement types và các cột tương ứng."""
    config = configparser.ConfigParser()
    if not os.path.exists(conf_file):
        raise FileNotFoundError(f"Config file not found: {conf_file}")
    
    config.read(conf_file)
    print(f"Sections found in .conf file: {config.sections()}")
    
    if 'common' not in config:
        raise KeyError("Section [common] not found in the .conf file. Please check the file format.")
    
    filter_line = config['common']['filter']
    measurement_types = [m.strip() for m in filter_line.split(',')]
    
    conf_data = {}
    for m_type in measurement_types:
        if m_type in config:
            columns = [col.strip() for col in config[m_type]['columns'].split(',')]
            conf_data[m_type] = columns
        else:
            print(f"Warning: Measurement type {m_type} not found in .conf file")
    
    return conf_data

def parse_time(time_str):
    """Chuyển đổi startTime từ XML sang định dạng TIME yêu cầu (YYYY-MM-DD HH:MM:SS)."""
    try:
        dt = datetime.fromisoformat(time_str.replace(':00', ''))
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception as e:
        raise ValueError(f"Error parsing time {time_str}: {str(e)}")

def extract_data(xml_file, conf_data):
    """Trích xuất dữ liệu từ file XML và trả về dictionary chứa dữ liệu cho từng measurement type."""
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        pm_setup = root.find('.//PMSetup')
        if pm_setup is None:
            print(f"No PMSetup found in {xml_file}")
            return {m_type: [] for m_type in conf_data.keys()}
        
        start_time = pm_setup.get('startTime')
        interval = pm_setup.get('interval')
        time_str = parse_time(start_time)
        
        data = {m_type: [] for m_type in conf_data.keys()}
        found_measurement_types = set()
        
        for pm_result in root.findall('.//PMMOResult'):
            mo = pm_result.find('MO')
            if mo is None:
                continue
            dn = mo.find('DN')
            if dn is None:
                continue
            key = dn.text
            
            pm_target = pm_result.find('PMTarget')
            if pm_target is None:
                continue
            measurement_type = pm_target.get('measurementType')
            found_measurement_types.add(measurement_type)
            
            if measurement_type in conf_data:
                row = {'KEY': key, 'TIME': time_str, 'RES': interval}
                for column in conf_data[measurement_type]:
                    value = pm_target.find(f".//{column}")
                    row[column] = value.text if value is not None else '0'
                data[measurement_type].append(row)
        
        print(f"Measurement types found in {xml_file}: {found_measurement_types}")
        return data
    except Exception as e:
        raise RuntimeError(f"Error extracting data from {xml_file}: {str(e)}")

def write_to_csv(all_data, output1_dir, output2_dir, folder_number, conf_data):
    """Ghi dữ liệu vào các file CSV trong output1 (thư mục con) và output2 (ghi đè)."""
    try:
        # Xử lý output1: Ghi vào thư mục con với số thứ tự
        output1_subdir = os.path.join(output1_dir, str(folder_number))
        if not os.path.exists(output1_subdir):
            os.makedirs(output1_subdir)
            print(f"Created output directory: {output1_subdir}")
        
        for m_type, rows in all_data.items():
            if not rows:
                print(f"No data to write for measurement type {m_type} in {output1_subdir}")
                continue
            csv_file = os.path.join(output1_subdir, f'{m_type}.csv')
            columns = ['TIME', 'RES', 'KEY'] + conf_data[m_type]
            
            with open(csv_file, 'a', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=columns, delimiter='|')
                for row in rows:
                    filtered_row = {k: v for k, v in row.items() if k in columns}
                    writer.writerow(filtered_row)
                print(f"Appended {len(rows)} rows to {csv_file}")
        
        # Xử lý output2: Ghi đè trực tiếp vào thư mục gốc với toàn bộ dữ liệu
        if not os.path.exists(output2_dir):
            os.makedirs(output2_dir)
            print(f"Created output directory: {output2_dir}")
        
        for m_type, rows in all_data.items():
            if not rows:
                print(f"No data to write for measurement type {m_type} in {output2_dir}")
                continue
            csv_file = os.path.join(output2_dir, f'{m_type}.csv')
            columns = ['TIME', 'RES', 'KEY'] + conf_data[m_type]
            
            with open(csv_file, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=columns, delimiter='|')
                for row in rows:
                    filtered_row = {k: v for k, v in row.items() if k in columns}
                    writer.writerow(filtered_row)
                print(f"Overwrote {len(rows)} rows to {csv_file}")
    except Exception as e:
        raise RuntimeError(f"Error writing to CSV: {str(e)}")

def decompress_gz(gz_file, temp_dir):
    """Giải nén file .xml.gz và trả về đường dẫn file .xml."""
    xml_file = os.path.join(temp_dir, os.path.basename(gz_file).replace('.gz', ''))
    try:
        with gzip.open(gz_file, 'rb') as f_in:
            with open(xml_file, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
        print(f"Decompressed {gz_file} to {xml_file}")
        return xml_file
    except Exception as e:
        raise RuntimeError(f"Error decompressing {gz_file}: {str(e)}")

def delete_temp_file(xml_file, temp_dir):
    """Xóa file tạm trong temp_dir nếu tồn tại."""
    if xml_file.startswith(temp_dir):
        try:
            os.remove(xml_file)
            print(f"Deleted temporary file {xml_file}")
        except Exception as e:
            print(f"Could not delete temporary file {xml_file}: {str(e)}")

def get_processed_files(log_file):
    """Lấy danh sách các file đã được xử lý từ file Excel log."""
    processed_files = set()
    if os.path.exists(log_file):
        wb = load_workbook(log_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2]:  # FILE_NAME column
                processed_files.add(row[2])  # Chỉ lấy tên file
        print(f"Processed files from log: {processed_files}")
    return processed_files

def main(input_dir, conf_file, output1_dir, output2_dir, temp_dir, log_file):
    """Xử lý tất cả các file .xml.gz và .xml trong thư mục."""
    create_excel_log(log_file)
    conf_data = parse_conf_file(conf_file)
    
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
        print(f"Created temporary directory: {temp_dir}")
    
    processed_files = get_processed_files(log_file)
    stt = 1  # Bắt đầu STT từ 1
    
    # Thu thập tất cả dữ liệu từ các file XML
    all_data = {m_type: [] for m_type in conf_data.keys()}
    folder_number = get_next_folder_number(output1_dir)
    
    # Xử lý file .xml.gz
    gz_files = glob.glob(os.path.join(input_dir, "*.xml.gz"))
    for gz_file in gz_files:
        xml_file = decompress_gz(gz_file, temp_dir)
        if os.path.basename(xml_file) in processed_files:
            print(f"Skipping {xml_file}: Already processed")
            delete_temp_file(xml_file, temp_dir)
            continue
        
        print(f"Processing {xml_file}")
        time_process = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        try:
            data = extract_data(xml_file, conf_data)
            for m_type, rows in data.items():
                all_data[m_type].extend(rows)
            
            append_to_excel_log(log_file, stt, os.path.dirname(xml_file), os.path.basename(xml_file), time_process, folder_number)
            print(f"Successfully processed {xml_file}")
            stt += 1
        except Exception as e:
            append_to_excel_log(log_file, stt, os.path.dirname(xml_file), os.path.basename(xml_file), time_process, folder_number)
            print(f"Failed to process {xml_file}: {str(e)}")
            stt += 1
        
        delete_temp_file(xml_file, temp_dir)
    
    # Xử lý file .xml
    xml_files = glob.glob(os.path.join(input_dir, "*.xml"))
    for xml_file in xml_files:
        if os.path.basename(xml_file) in processed_files:
            print(f"Skipping {xml_file}: Already processed")
            continue
        
        print(f"Processing {xml_file}")
        time_process = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        try:
            data = extract_data(xml_file, conf_data)
            for m_type, rows in data.items():
                all_data[m_type].extend(rows)
            
            append_to_excel_log(log_file, stt, os.path.dirname(xml_file), os.path.basename(xml_file), time_process, folder_number)
            print(f"Successfully processed {xml_file}")
            stt += 1
        except Exception as e:
            append_to_excel_log(log_file, stt, os.path.dirname(xml_file), os.path.basename(xml_file), time_process, folder_number)
            print(f"Failed to process {xml_file}: {str(e)}")
            stt += 1
    
    # Ghi tất cả dữ liệu vào output1 và output2
    if any(all_data[m_type] for m_type in all_data):
        write_to_csv(all_data, output1_dir, output2_dir, folder_number, conf_data)

if __name__ == '__main__':
    input_dir = r'/opt/airflow/data/LTE-Nokia-MB_10.51.144.12/source'
    conf_file = r'/opt/airflow/data/LTE-Nokia-MB_10.51.144.12/nsnxml2csv2.conf'
    output1_dir = r'/opt/airflow/data/LTE-Nokia-MB_10.51.144.12/archive'
    output2_dir = r'/opt/airflow/data/LTE-Nokia-MB_10.51.144.12/output'
    temp_dir = r'/opt/airflow/data/LTE-Nokia-MB_10.51.144.12/staging'
    log_file = r'/opt/airflow/data/LTE-Nokia-MB_10.51.144.12/logs/extraction_log.xlsx'
    
    main(input_dir, conf_file, output1_dir, output2_dir, temp_dir, log_file)